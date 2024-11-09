# Copyright (c) 2024, TnologiaRD and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
import frappe
from frappe.utils import flt
from frappe.model.document import Document
from frappe.utils.csvutils import UnicodeWriter
from datetime import datetime
import time
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side, NamedStyle


class Reporte607(Document):
	pass


@frappe.whitelist()
def validate_pending_invoices(from_date, to_date):

    # Recuperar los valores de los campos del Doctype DGII Reports Settings
    settings = frappe.get_single("DGII Reports Settings")
    required_fields = [
    settings.itbis_facturado_607,
    settings.itbis_percibido_607,
    settings.isr_percibido_607,
    settings.isc_607,
    settings.propina_legal_607,
    settings.ret_607_itbis_retenido_por_terceros,
    settings.ret_607_retencion_renta_por_terceros,
    settings.otros_impuestos_tasas_607
    ]

    # Verificar que todos los campos tengan un valor
    if not all(required_fields): 
        settings_url = frappe.utils.get_url_to_form("DGII Reports Settings", "DGII Reports Settings") + "#dgii-reports-settings-cuentas_607_tab"
        frappe.log_error("DGII Reports Setting Missing Accounts", "validate_missing_accounts")
        return {"message": frappe._("Se deben configurar las cuentas para el reporte 607 en el documento <a href='{0}'>DGII Reports Settings</a>").format(settings_url)}


    draft_invoices = frappe.db.sql("""
        SELECT 
            COUNT(*) 
        FROM 
            `tabSales Invoice` 
        WHERE 
            docstatus = 0 AND posting_date BETWEEN '%s' AND '%s'
    """ % (from_date, to_date))

    if draft_invoices[0][0] > 0:
        frappe.log_error("Facturas pendientes encontradas", "validate_pending_invoices")
        return {"message": "Hay facturas de venta pendientes de procesar. Por favor, complete las facturas pendientes antes de generar el reporte."}
    
    frappe.log_error("No hay facturas pendientes", "validate_pending_invoices")
    return {"message": ""}


@frappe.whitelist()
def get_file_address_1(from_date, to_date, decimal_places=2):
    # Obtener el tax_id del doctype Company
    company = frappe.get_doc("Company", frappe.defaults.get_user_default("Company"))
    rnc = company.tax_id.replace("-", "") if company.tax_id else ""

    # Calcular el periodo
    from_date_obj = datetime.strptime(from_date, "%Y-%m-%d")
    to_date_obj = datetime.strptime(to_date, "%Y-%m-%d")
    periodo = f"{from_date_obj.year}{min(from_date_obj.month, to_date_obj.month):02d}"

    # Consulta SQL para obtener los datos necesarios
    query = """
    SELECT 
        cust.tax_id, 
        cust.custom_tipo_rnc as tipo_identificacion,
        sinv.custom_ncf as ncf, 
        sinv.custom_return_against_ncf as ncf_modificado,
        tipo_ing.codigo as tipo_de_ingreso,  -- Recuperar la columna codigo de la tabla tabTipo de Ingreso
        sinv.posting_date as fecha_comprobante, 
        sinv.due_date as fecha_retencion,
        sinv.base_total as monto_facturado, 
        sinv.base_total_taxes_and_charges as itbis_facturado,
        '' as itbis_retenido_terceros,  -- No disponible en la tabla
        '' as itbis_percibido,  -- No disponible en la tabla
        '' as retencion_renta_terceros,  -- No disponible en la tabla
        '' as isr_percibido,  -- No disponible en la tabla
        '' as impuesto_selectivo_consumo,  -- No disponible en la tabla
        '' as otros_impuestos_tasas,  -- No disponible en la tabla
        '' as monto_propina_legal,  -- No disponible en la tabla
        CASE WHEN COALESCE(SUM(CASE WHEN pe.mode_of_payment = 'Efectivo' THEN pe.paid_amount ELSE 0 END), 0) = 0 THEN '' ELSE COALESCE(SUM(CASE WHEN pe.mode_of_payment = 'Efectivo' THEN pe.paid_amount ELSE 0 END), 0) END as efectivo,  -- Calcular el total de pagos con Modo de Pago Efectivo
        CASE WHEN COALESCE(SUM(CASE WHEN pe.mode_of_payment = 'Transferencia bancaria' THEN pe.paid_amount ELSE 0 END), 0) = 0 THEN '' ELSE COALESCE(SUM(CASE WHEN pe.mode_of_payment = 'Transferencia bancaria' THEN pe.paid_amount ELSE 0 END), 0) END as cheque_transferencia_deposito,  -- Calcular el total de pagos con Modo de Pago Transferencia bancaria
        CASE WHEN COALESCE(SUM(CASE WHEN pe.mode_of_payment = 'Tarjetas de credito' THEN pe.paid_amount ELSE 0 END), 0) = 0 THEN '' ELSE COALESCE(SUM(CASE WHEN pe.mode_of_payment = 'Tarjetas de credito' THEN pe.paid_amount ELSE 0 END), 0) END as tarjeta_debito_credito,  -- Calcular el total de pagos con Modo de Pago Tarjetas de credito
        '' as venta_credito,  -- No disponible en la tabla
        '' as bonos_certificados_regalo,  -- No disponible en la tabla
        '' as permuta,  -- No disponible en la tabla
        '' as otras_formas_ventas  -- No disponible en la tabla
    FROM 
        `tabSales Invoice` AS sinv 
    JOIN 
        tabCustomer AS cust on sinv.customer = cust.name 
    LEFT JOIN 
        `tabPayment Entry Reference` AS per ON per.reference_name = sinv.name
    LEFT JOIN 
        `tabPayment Entry` AS pe ON pe.name = per.parent
    LEFT JOIN 
        `tabTipo de Ingreso` AS tipo_ing ON sinv.custom_tipo_de_ingreso = tipo_ing.tipo_de_ingreso
    WHERE 
        sinv.custom_ncf NOT LIKE 'SINV-%%' AND sinv.docstatus = 1 AND sinv.posting_date BETWEEN %(from_date)s AND %(to_date)s
    GROUP BY 
        sinv.name
    """

    # Ejecutar la consulta
    result = frappe.db.sql(query, {"from_date": from_date, "to_date": to_date}, as_dict=True)

    # Número de registros
    numero_registros = len(result)

    # Crear el archivo en memoria usando UnicodeWriter
    w = UnicodeWriter()

    # Agregar la primera línea
    w.writerow([
        "607", 
        rnc, 
        periodo, 
        numero_registros
    ])

    def format_amount(amount):
        if amount == '':
            return amount
        try:
            amount = abs(float(amount))  # Convertir a valor absoluto
        except ValueError:
            return amount
        if amount == int(amount):
            return str(int(amount))
        else:
            return f"{amount:.{decimal_places}f}"
        
    for row in result:
        if row.tipo_identificacion == "RNC":
            tipo_identificacion = "1"
        elif row.tipo_identificacion == "Cédula":
            tipo_identificacion = "2"
        elif row.tipo_identificacion == "Pasaporte":
            tipo_identificacion = "3"
        else:
            tipo_identificacion = "2"  # Valor por defecto si no coincide con ninguno

        fecha_retencion = row.fecha_retencion.strftime("%Y%m%d") if row.fecha_retencion and row.retencion_renta_terceros not in ["", 0] else ""
        w.writerow([
            row.tax_id.replace("-", "") if row.tax_id else "", 
            tipo_identificacion, 
            row.ncf, 
            row.ncf_modificado, 
            int(row.tipo_de_ingreso),  # Convertir tipo_de_ingreso a entero
            row.fecha_comprobante.strftime("%Y%m%d") if row.fecha_comprobante else "", 
            fecha_retencion, 
            format_amount(row.monto_facturado), 
            format_amount(row.itbis_facturado), 
            row.itbis_retenido_terceros, 
            row.itbis_percibido, 
            row.retencion_renta_terceros, 
            row.isr_percibido, 
            row.impuesto_selectivo_consumo, 
            row.otros_impuestos_tasas, 
            row.monto_propina_legal, 
            format_amount(row.efectivo) if row.efectivo != 0 else "", 
            format_amount(row.cheque_transferencia_deposito) if row.cheque_transferencia_deposito != 0 else "", 
            format_amount(row.tarjeta_debito_credito) if row.tarjeta_debito_credito != 0 else "",
            row.venta_credito, 
            row.bonos_certificados_regalo, 
            row.permuta, 
            row.otras_formas_ventas
        ])

    # Convertir el contenido a texto con delimitador de pipes
    content = w.getvalue().replace(",", "|").replace('"', '')

    # Devolver el contenido del archivo de texto como descarga
    frappe.response['filename'] = "Reporte_607_%s.txt" % time.strftime("%Y%m%d_%H%M%S")
    frappe.response['filecontent'] = content
    frappe.response['type'] = 'download'

def format_amount(amount, decimal_places=2):
    if amount == '' or amount == 0 or amount == '0.000000000':
        return ''
    try:
        amount = abs(float(amount))  # Convertir a valor absoluto
    except ValueError:
        return amount
    if amount == int(amount):
        return str(int(amount))
    else:
        return f"{amount:.{decimal_places}f}"

def filter_results(results, filter_condition):
    """Filtrar los resultados basados en la condición dada."""
    return [row for row in results if filter_condition(row)]



def generate_sheet(ws, data, rnc, periodo, numero_registros, include_totals=False, total_columns=None):
    """Generar una hoja de Excel con los datos proporcionados."""
    # Estilos
    header_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")  # Verde oscuro
    header_font = Font(name="Tahoma", size=11, color="FFFFFF")
    detail_font = Font(name="Tahoma", size=9, color="FFFFFF", bold=True)
    header_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    detail_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    text_style = NamedStyle(name="text_style", number_format="@")

    # Agregar la primera línea
    ws.merge_cells('A1:B1')
    ws['A1'] = "RNC o Cédula"
    ws['C1'] = rnc
    ws['A1'].fill = header_fill
    ws['A1'].font = header_font
    ws['A1'].border = header_border
    ws['B1'].border = header_border
    ws['C1'].border = header_border  # Aplicar borde a la celda C1
    ws['A1'].alignment = Alignment(horizontal='right')
    ws['C1'].alignment = Alignment(horizontal='right')

    # Agregar la segunda línea
    ws.merge_cells('A2:B2')
    ws['A2'] = "Periodo"
    ws['C2'] = periodo
    ws['A2'].fill = header_fill
    ws['A2'].font = header_font
    ws['A2'].border = header_border
    ws['B2'].border = header_border
    ws['C2'].border = header_border
    ws['A2'].alignment = Alignment(horizontal='right')
    ws['C2'].alignment = Alignment(horizontal='right')

    # Agregar la tercera línea
    ws.merge_cells('A3:B3')
    ws['A3'] = "Cantidad Registros"
    ws['C3'] = numero_registros
    ws['A3'].fill = header_fill
    ws['A3'].font = header_font
    ws['A3'].border = header_border
    ws['B3'].border = header_border
    ws['C3'].border = header_border
    ws['A3'].alignment = Alignment(horizontal='right')
    ws['C3'].alignment = Alignment(horizontal='right')

    # Dejar un renglón en blanco
    ws.append([])

    # Agregar el encabezado de detalles
    ws.merge_cells('B5:X5')  # Combinar hasta la columna X5
    ws['B5'] = "Detalles"
    ws['B5'].alignment = Alignment(horizontal='center')
    ws['B5'].fill = header_fill
    ws['B5'].font = detail_font
    ws['B5'].border = header_border

    # Agregar la fila de números de columna
    column_numbers = ["", 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23]
    ws.append(column_numbers)
    for col_num in range(2, 25):
        cell = ws.cell(row=6, column=col_num)
        cell.fill = header_fill
        cell.font = detail_font
        cell.border = header_border
        cell.alignment = Alignment(horizontal='center')  # Centrar el texto

    # Agregar el encabezado de las columnas
    headers = [
        "No",
        "RNC/Cédula o Pasaporte",
        "Tipo Identificación",
        "Número Comprobante Fiscal",
        "Número Comprobante Fiscal Modificado",
        "Tipo de Ingreso",
        "Fecha Comprobante",
        "Fecha de Retención",
        "Monto Facturado",
        "ITBIS Facturado",
        "ITBIS Retenido por Terceros",
        "ITBIS Percibido",
        "Retención Renta por Terceros",
        "ISR Percibido",
        "Impuesto Selectivo al Consumo",
        "Otros Impuestos/Tasas",
        "Monto Propina Legal",
        "Efectivo",
        "Cheque/ Transferencia/ Depósito",
        "Tarjeta Débito/Crédito",
        "Venta a Crédito",
        "Bonos o Certificados de Regalo",
        "Permuta",
        "Otras Formas de Ventas"
    ]
    ws.append(headers)
    for col_num in range(1, 25):
        cell = ws.cell(row=7, column=col_num)
        cell.fill = header_fill
        cell.font = detail_font
        cell.border = header_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Centrar y ajustar el texto

    # Ajustar el ancho de las columnas
    column_widths = {
        'A': 10,
        'B': 18,
        'C': 22,
        'D': 25,
        'E': 28,
        'F': 40,
        'G': 18,
        'H': 18,
        'I': 18,
        'J': 18,
        'K': 18,
        'L': 18,
        'M': 18,
        'N': 18,
        'O': 18,
        'P': 18,
        'Q': 18,
        'R': 18,
        'S': 18,
        'T': 18,
        'U': 18,
        'V': 18,
        'W': 18,
        'X': 18
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Llenar las columnas del reporte 607 con los datos obtenidos
    for idx, row in enumerate(data, start=1):
        if row['Tipo Identificación'] == "RNC":
            tipo_identificacion = "1"
        elif row['Tipo Identificación'] == "Cédula":
            tipo_identificacion = "2"
        elif row['Tipo Identificación'] == "Pasaporte":
            tipo_identificacion = "3"
        else:
            tipo_identificacion = "2"  # Valor por defecto si no coincide con ninguno

        fecha_retencion = ""
        if row['Fecha de Retención'] and (
            row['ITBIS Retenido por Terceros'] not in ["", 0] or 
            row['Retención Renta por Terceros'] not in ["", 0]
        ):
            fecha_retencion = row['Fecha de Retención'].strftime("%Y%m%d")
        ws.append([
            idx,
            row['RNC/Cédula o Pasaporte'].replace("-", "") if row['RNC/Cédula o Pasaporte'] else '',
            tipo_identificacion,
            row['Número Comprobante Fiscal'] or '',
            row['Número Comprobante Fiscal Modificado'] or '',
            row['Tipo de Ingreso'] if row['Tipo de Ingreso'] else '',
            row['Fecha Comprobante'].strftime("%Y%m%d") if row['Fecha Comprobante'] else '',
            fecha_retencion,
            format_amount(row['Monto Facturado']),
            format_amount(row['ITBIS Facturado']),
            format_amount(row['ITBIS Retenido por Terceros']) or '',
            format_amount(row['ITBIS Percibido']) or '',
            format_amount(row['Retención Renta por Terceros']) or '',
            format_amount(row['ISR Percibido']) or '',
            format_amount(row['Impuesto Selectivo al Consumo']) or '',
            format_amount(row['Otros Impuestos/Tasas']) or '',
            format_amount(row['Monto Propina Legal']) or '',
            format_amount(row['Efectivo']),
            format_amount(row['Cheque/ Transferencia/ Depósito']),
            format_amount(row['Tarjeta Débito/Crédito']),
            format_amount(row['Venta a Crédito']) or '',
            format_amount(row['Bonos o Certificados de Regalo']) or '',
            format_amount(row['Permuta']) or '',
            format_amount(row['Otras Formas de Ventas']) or ''
        ])
        if "text_style" not in ws.parent.named_styles:
            ws.parent.add_named_style(text_style)
        for col_num in range(1, 25):
            cell = ws.cell(row=idx + 7, column=col_num)
            cell.style = text_style  # Aplicar el estilo de texto plano
            if col_num == 1:
                cell.fill = detail_fill
                cell.alignment = Alignment(horizontal='center')  # Centrar el texto horizontalmente

    # Incluir totales si es necesario
    if include_totals and total_columns:
        total_row = [""] * 24
        total_row[0] = "Totales"
        for col in total_columns:
            col_letter = chr(65 + col)  # Convertir índice de columna a letra
            total_formula = f"=SUMPRODUCT({col_letter}8:{col_letter}{len(data) + 7}*1)"
            total_row[col] = total_formula
        ws.append(total_row)

        # Aplicar estilo a las celdas de totales
        for col in total_columns:
            col_letter = chr(65 + col)  # Convertir índice de columna a letra
            cell = ws.cell(row=len(data) + 8, column=col + 1)
            cell.number_format = '0.00'  # Formato numérico
            cell.alignment = Alignment(horizontal='left', vertical='center')  # Alinear a la izquierda
            cell.font = Font(bold=True)  # Poner en negritas
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gris claro
            cell.value = f'=IF(SUMPRODUCT({col_letter}8:{col_letter}{len(data) + 7}*1)=0, "", TEXT(SUMPRODUCT({col_letter}8:{col_letter}{len(data) + 7}*1), "0.00"))'


@frappe.whitelist()
def get_file_address(from_date, to_date, decimal_places=2, include_totals=False, total_columns=None):
    # Obtener el tax_id del doctype Company
    company = frappe.get_doc("Company", frappe.defaults.get_user_default("Company"))
    rnc = company.tax_id.replace("-", "") if company.tax_id else ""

    # Calcular el periodo
    from_date_obj = datetime.strptime(from_date, "%Y-%m-%d")
    to_date_obj = datetime.strptime(to_date, "%Y-%m-%d")
    periodo = f"{from_date_obj.year}{min(from_date_obj.month, to_date_obj.month):02d}"

    # Recuperar los nombres de las cuentas del Doctype DGII Reports Settings
    settings = frappe.get_single("DGII Reports Settings")
    itbis_facturado = settings.itbis_facturado_607 or ''
    itbis_percibido = settings.itbis_percibido_607 or ''
    isr_percibido = settings.isr_percibido_607 or ''
    impuesto_selectivo_al_consumo = settings.isc_607 or ''
    propina_legal = settings.propina_legal_607 or ''
    itbis_retenido_por_terceros = settings.ret_607_itbis_retenido_por_terceros or ''
    retencion_renta_por_terceros = settings.ret_607_retencion_renta_por_terceros or ''

    # Obtener las cuentas de la tabla otros_impuestos
    otros_impuestos_cuentas = [frappe.db.escape(item.cuenta) for item in settings.otros_impuestos_tasas_607]

    # Construir las condiciones SQL para otros_impuestos
    if otros_impuestos_cuentas:
        otros_impuestos_condition_ptc = " OR ".join([f"ptc.account_head = {cuenta}" for cuenta in otros_impuestos_cuentas])
        otros_impuestos_sql_ptc = f"SUM(CASE WHEN {otros_impuestos_condition_ptc} THEN ptc.tax_amount ELSE 0 END) AS `Otros Impuestos/Tasas`"
    else:
        otros_impuestos_sql_ptc = "0 AS `Otros Impuestos/Tasas`"

    efectivo = "01 - EFECTIVO"
    cheque_transferencia_deposito = "02 - CHEQUES/TRANSFERENCIAS/DEPÓSITO"
    tarjeta_debito_credito = "03 - TARJETA CRÉDITO/DÉBITO"
    permuta = "05 - PERMUTA"
    venta_credito = "08 - VENTA CREDITO"
    bonos_certificados_regalo = "09 - BONOS O CERTIFICADOS DE REGALO"
    otras_formas_ventas = "10 - OTRAS FORMAS DE VENTAS"

    # Consulta SQL para obtener los datos necesarios
    query = f"""
    SELECT cust.tax_id AS `RNC/Cédula o Pasaporte`, cust.custom_tipo_rnc AS `Tipo Identificación`,
    sinv.custom_ncf AS `Número Comprobante Fiscal`, sinv.custom_return_against_ncf AS `Número Comprobante Fiscal Modificado`,
    CONCAT(tipo_ing.codigo, ' - ', tipo_ing.tipo_de_ingreso) AS `Tipo de Ingreso`, sinv.posting_date AS `Fecha Comprobante`,
    sinv.due_date AS `Fecha de Retención`, sinv.base_total AS `Monto Facturado`,
    SUM(CASE WHEN ptc.account_head = '{itbis_facturado}' THEN ptc.tax_amount ELSE 0 END) AS `ITBIS Facturado`,
    SUM(CASE WHEN ptc.account_head = '{itbis_retenido_por_terceros}' THEN ptc.tax_amount ELSE 0 END) AS `ITBIS Retenido por Terceros`,
    SUM(CASE WHEN ptc.account_head = '{itbis_percibido}' THEN ptc.tax_amount ELSE 0 END) AS `ITBIS Percibido`,
    SUM(CASE WHEN ptc.account_head = '{retencion_renta_por_terceros}' THEN ptc.tax_amount ELSE 0 END) AS `Retención Renta por Terceros`,
    SUM(CASE WHEN ptc.account_head = '{isr_percibido}' THEN ptc.tax_amount ELSE 0 END) AS `ISR Percibido`,
    SUM(CASE WHEN ptc.account_head = '{impuesto_selectivo_al_consumo}' THEN ptc.tax_amount ELSE 0 END) AS `Impuesto Selectivo al Consumo`,
    {otros_impuestos_sql_ptc},
    SUM(CASE WHEN ptc.account_head = '{propina_legal}' THEN ptc.tax_amount ELSE 0 END) AS `Monto Propina Legal`,
    CASE WHEN COALESCE(SUM(CASE WHEN LOWER(mop.custom_dgii_mode_of_payment) = '{efectivo}' THEN pe.paid_amount ELSE 0 END), 0) = 0 THEN '' ELSE COALESCE(SUM(CASE WHEN LOWER(mop.custom_dgii_mode_of_payment) = '{efectivo}' THEN pe.paid_amount ELSE 0 END), 0) END AS `Efectivo`,
    CASE WHEN COALESCE(SUM(CASE WHEN LOWER(mop.custom_dgii_mode_of_payment) = '{cheque_transferencia_deposito}' THEN pe.paid_amount ELSE 0 END), 0) = 0 THEN '' ELSE COALESCE(SUM(CASE WHEN LOWER(mop.custom_dgii_mode_of_payment) = '{cheque_transferencia_deposito}' THEN pe.paid_amount ELSE 0 END), 0) END AS `Cheque/ Transferencia/ Depósito`,
    CASE WHEN COALESCE(SUM(CASE WHEN LOWER(mop.custom_dgii_mode_of_payment) = '{tarjeta_debito_credito}' THEN pe.paid_amount ELSE 0 END), 0) = 0 THEN '' ELSE COALESCE(SUM(CASE WHEN LOWER(mop.custom_dgii_mode_of_payment) = '{tarjeta_debito_credito}' THEN pe.paid_amount ELSE 0 END), 0) END AS `Tarjeta Débito/Crédito`,
    CASE WHEN COALESCE(SUM(CASE WHEN LOWER(mop.custom_dgii_mode_of_payment) = '{venta_credito}' THEN pe.paid_amount ELSE 0 END), 0) = 0 THEN '' ELSE COALESCE(SUM(CASE WHEN LOWER(mop.custom_dgii_mode_of_payment) = '{venta_credito}' THEN pe.paid_amount ELSE 0 END), 0) END AS `Venta a Crédito`,
    CASE WHEN COALESCE(SUM(CASE WHEN LOWER(mop.custom_dgii_mode_of_payment) = '{bonos_certificados_regalo}' THEN pe.paid_amount ELSE 0 END), 0) = 0 THEN '' ELSE COALESCE(SUM(CASE WHEN LOWER(mop.custom_dgii_mode_of_payment) = '{bonos_certificados_regalo}' THEN pe.paid_amount ELSE 0 END), 0) END AS `Bonos o Certificados de Regalo`,
    CASE WHEN COALESCE(SUM(CASE WHEN LOWER(mop.custom_dgii_mode_of_payment) = '{permuta}' THEN pe.paid_amount ELSE 0 END), 0) = 0 THEN '' ELSE COALESCE(SUM(CASE WHEN LOWER(mop.custom_dgii_mode_of_payment) = '{permuta}' THEN pe.paid_amount ELSE 0 END), 0) END AS `Permuta`,
    CASE WHEN COALESCE(SUM(CASE WHEN LOWER(mop.custom_dgii_mode_of_payment) = '{otras_formas_ventas}' THEN pe.paid_amount ELSE 0 END), 0) = 0 THEN '' ELSE COALESCE(SUM(CASE WHEN LOWER(mop.custom_dgii_mode_of_payment) = '{otras_formas_ventas}' THEN pe.paid_amount ELSE 0 END), 0) END AS `Otras Formas de Ventas`
    FROM `tabSales Invoice` AS sinv 
    JOIN tabCustomer AS cust ON sinv.customer = cust.name 
    LEFT JOIN `tabPayment Entry Reference` AS per ON per.reference_name = sinv.name
    LEFT JOIN `tabPayment Entry` AS pe ON pe.name = per.parent
    LEFT JOIN `tabMode of Payment` AS mop ON pe.mode_of_payment = mop.name
    LEFT JOIN `tabTipo de Ingreso` AS tipo_ing ON sinv.custom_tipo_de_ingreso = tipo_ing.tipo_de_ingreso
    LEFT JOIN `tabSales Taxes and Charges` AS ptc ON sinv.name = ptc.parent
    WHERE sinv.custom_ncf NOT LIKE 'SINV-%%' AND sinv.docstatus = 1 AND sinv.posting_date BETWEEN %(from_date)s AND %(to_date)s
    GROUP BY sinv.name

    UNION

    SELECT cust.tax_id AS `RNC/Cédula o Pasaporte`, cust.custom_tipo_rnc AS `Tipo Identificación`,
    sinv.custom_ncf AS `Número Comprobante Fiscal`, sinv.custom_return_against_ncf AS `Número Comprobante Fiscal Modificado`,
    CONCAT(tipo_ing.codigo, ' - ', tipo_ing.tipo_de_ingreso) AS `Tipo de Ingreso`, sinv.posting_date AS `Fecha Comprobante`,
    sinv.due_date AS `Fecha de Retención`, sinv.base_total AS `Monto Facturado`,
    SUM(CASE WHEN ptc.account_head = '{itbis_facturado}' THEN ptc.tax_amount ELSE 0 END) AS `ITBIS Facturado`,
    SUM(CASE WHEN ptc.account_head = '{itbis_retenido_por_terceros}' THEN ptc.tax_amount ELSE 0 END) AS `ITBIS Retenido por Terceros`,
    SUM(CASE WHEN ptc.account_head = '{itbis_percibido}' THEN ptc.tax_amount ELSE 0 END) AS `ITBIS Percibido`,
    SUM(CASE WHEN ptc.account_head = '{retencion_renta_por_terceros}' THEN ptc.tax_amount ELSE 0 END) AS `Retención Renta por Terceros`,
    SUM(CASE WHEN ptc.account_head = '{isr_percibido}' THEN ptc.tax_amount ELSE 0 END) AS `ISR Percibido`,
    SUM(CASE WHEN ptc.account_head = '{impuesto_selectivo_al_consumo}' THEN ptc.tax_amount ELSE 0 END) AS `Impuesto Selectivo al Consumo`,
    {otros_impuestos_sql_ptc},
    SUM(CASE WHEN ptc.account_head = '{propina_legal}' THEN ptc.tax_amount ELSE 0 END) AS `Monto Propina Legal`,
    CASE WHEN COALESCE(SUM(CASE WHEN LOWER(mop.custom_dgii_mode_of_payment) = '{efectivo}' THEN pe.paid_amount ELSE 0 END), 0) = 0 THEN '' ELSE COALESCE(SUM(CASE WHEN LOWER(mop.custom_dgii_mode_of_payment) = '{efectivo}' THEN pe.paid_amount ELSE 0 END), 0) END AS `Efectivo`,
    CASE WHEN COALESCE(SUM(CASE WHEN LOWER(mop.custom_dgii_mode_of_payment) = '{cheque_transferencia_deposito}' THEN pe.paid_amount ELSE 0 END), 0) = 0 THEN '' ELSE COALESCE(SUM(CASE WHEN LOWER(mop.custom_dgii_mode_of_payment) = '{cheque_transferencia_deposito}' THEN pe.paid_amount ELSE 0 END), 0) END AS `Cheque/ Transferencia/ Depósito`,
    CASE WHEN COALESCE(SUM(CASE WHEN LOWER(mop.custom_dgii_mode_of_payment) = '{tarjeta_debito_credito}' THEN pe.paid_amount ELSE 0 END), 0) = 0 THEN '' ELSE COALESCE(SUM(CASE WHEN LOWER(mop.custom_dgii_mode_of_payment) = '{tarjeta_debito_credito}' THEN pe.paid_amount ELSE 0 END), 0) END AS `Tarjeta Débito/Crédito`,
    CASE WHEN COALESCE(SUM(CASE WHEN LOWER(mop.custom_dgii_mode_of_payment) = '{venta_credito}' THEN pe.paid_amount ELSE 0 END), 0) = 0 THEN '' ELSE COALESCE(SUM(CASE WHEN LOWER(mop.custom_dgii_mode_of_payment) = '{venta_credito}' THEN pe.paid_amount ELSE 0 END), 0) END AS `Venta a Crédito`,
    CASE WHEN COALESCE(SUM(CASE WHEN LOWER(mop.custom_dgii_mode_of_payment) = '{bonos_certificados_regalo}' THEN pe.paid_amount ELSE 0 END), 0) = 0 THEN '' ELSE COALESCE(SUM(CASE WHEN LOWER(mop.custom_dgii_mode_of_payment) = '{bonos_certificados_regalo}' THEN pe.paid_amount ELSE 0 END), 0) END AS `Bonos o Certificados de Regalo`,
    CASE WHEN COALESCE(SUM(CASE WHEN LOWER(mop.custom_dgii_mode_of_payment) = '{permuta}' THEN pe.paid_amount ELSE 0 END), 0) = 0 THEN '' ELSE COALESCE(SUM(CASE WHEN LOWER(mop.custom_dgii_mode_of_payment) = '{permuta}' THEN pe.paid_amount ELSE 0 END), 0) END AS `Permuta`,
    CASE WHEN COALESCE(SUM(CASE WHEN LOWER(mop.custom_dgii_mode_of_payment) = '{otras_formas_ventas}' THEN pe.paid_amount ELSE 0 END), 0) = 0 THEN '' ELSE COALESCE(SUM(CASE WHEN LOWER(mop.custom_dgii_mode_of_payment) = '{otras_formas_ventas}' THEN pe.paid_amount ELSE 0 END), 0) END AS `Otras Formas de Ventas`
    FROM `tabSales Invoice` AS sinv 
    JOIN tabCustomer AS cust ON sinv.customer = cust.name 
    LEFT JOIN `tabPayment Entry Reference` AS per ON per.reference_name = sinv.name
    LEFT JOIN `tabPayment Entry` AS pe ON pe.name = per.parent
    LEFT JOIN `tabMode of Payment` AS mop ON pe.mode_of_payment = mop.name
    LEFT JOIN `tabTipo de Ingreso` AS tipo_ing ON sinv.custom_tipo_de_ingreso = tipo_ing.tipo_de_ingreso
    LEFT JOIN `tabSales Taxes and Charges` AS ptc ON sinv.name = ptc.parent
    WHERE sinv.custom_ncf NOT LIKE 'SINV-%%' AND sinv.docstatus = 1 AND sinv.outstanding_amount = 0
    AND sinv.posting_date < %(from_date)s AND pe.posting_date BETWEEN %(from_date)s AND %(to_date)s AND pe.docstatus = 1
    GROUP BY sinv.name
    """

    # Ejecutar la consulta
    results = frappe.db.sql(query, {"from_date": from_date, "to_date": to_date}, as_dict=True)

    # Número de registros
    numero_registros = len(results)

    # Filtrar los resultados
    all_except_b02 = filter_results(results, lambda row: not row['Número Comprobante Fiscal'].startswith('B02'))
    only_b02 = filter_results(results, lambda row: row['Número Comprobante Fiscal'].startswith('B02'))

    # Crear el archivo Excel en memoria
    output = BytesIO()
    wb = Workbook()

    # Eliminar la hoja por defecto
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Generar la hoja con todos los datos excepto los que inician con "B02"
    ws_all_except_b02 = wb.create_sheet(title="607")
    generate_sheet(ws_all_except_b02, all_except_b02, rnc, periodo, len(all_except_b02), include_totals, total_columns)

    # Generar la hoja con solo los datos que inician con "B02"
    ws_only_b02 = wb.create_sheet(title="B02")
    generate_sheet(ws_only_b02, only_b02, rnc, periodo, len(only_b02), include_totals=True, total_columns=[8, 9, 10,11,12,13,14,15,16,17,18,19,20,21,22])

    # Guardar el archivo Excel en memoria
    wb.save(output)
    output.seek(0)

    # Devolver el contenido del archivo Excel como descarga
    frappe.response['filename'] = "Reporte_607_%s.xlsx" % time.strftime("%Y%m%d_%H%M%S")
    frappe.response['filecontent'] = output.getvalue()
    frappe.response['type'] = 'download'