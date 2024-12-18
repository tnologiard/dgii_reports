# Copyright (c) 2024, TnologiaRD and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, cint, flt, formatdate, format_datetime
from frappe.model.document import Document
from frappe.utils.csvutils import UnicodeWriter
from datetime import datetime, date
import time
import os
from frappe import _
from io import StringIO
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side, NamedStyle


class Reporte606(Document):
	pass

@frappe.whitelist()
def validate_pending_invoices(from_date, to_date):

    # # Recuperar los valores de los campos del Doctype DGII Reports Settings
    # settings = frappe.get_single("DGII Reports Settings")
    # required_fields = [
    #     settings.itbis_facturado,
    #     settings.ret606_itbis_retenido,
    #     settings.isc,
    #     settings.propina_legal,
    #     settings.ret606_isr,
    #     settings.otros_impuestos,
    #     # settings.itbis_proporcionalidad,
    #     # settings.itbis_costo,
    # ]

    # # Verificar que todos los campos tengan un valor
    # if not all(required_fields):
    #     # frappe.log_error("DGII Reports Setting Missing Accounts", "validate_missing_accounts")
    #     # return {"message": frappe._(""Todas las cuentas son requeridas excepto por ahora itbis_costo y itbis_proporcionalidad para el reporte 606 en el documento DGII Reports Settings. en el documento <a href='{0}'>DGII Reports Settings</a>").format(settings_url)} "}
    #     settings_url = frappe.utils.get_url_to_form("DGII Reports Settings", "DGII Reports Settings") 
    #     frappe.log_error("DGII Reports Setting Missing Accounts", "validate_missing_accounts")
    #     return {"message": frappe._("Se deben configurar todas las cuentas excepto por ahora itbis_costo y itbis_proporcionalidad para el reporte 606 en el documento <a href='{0}'>DGII Reports Settings</a>").format(settings_url)}

    # Consulta para contar facturas en borrador
    draft_invoices = frappe.db.sql("""
        SELECT 
            COUNT(*) 
        FROM 
            `tabPurchase Invoice` 
        WHERE 
            docstatus = 0 AND posting_date BETWEEN '%s' AND '%s'
    """ % (from_date, to_date))

    if draft_invoices[0][0] > 0:
        frappe.log_error("Facturas pendientes encontradas", "validate_pending_invoices")
        return {"message": "Hay facturas de compra pendientes de procesar. Por favor, complete las facturas pendientes antes de generar el reporte."}
    
    pending_invoices = frappe.db.sql("""
        SELECT 
            COUNT(*) 
        FROM 
            `tabPurchase Invoice` 
        WHERE 
            docstatus = 1 AND outstanding_amount > 0 AND posting_date BETWEEN '%s' AND '%s' AND (bill_no LIKE 'B13%%' OR bill_no LIKE 'B11%%')
    """ % (from_date, to_date))

    if pending_invoices[0][0] > 0:
        frappe.log_error("Comprobantes de gastos menores o compras no pagadas encontradas", "validate_pending_invoices")
        return {"message": "Hay comprobantes de gastos menores (B13) o comprobantes de compras (B11) no pagados. Por favor, complete el pago de estas facturas antes de generar el reporte."}
    
    frappe.log_error("No hay facturas pendientes", "validate_pending_invoices")
    return {"message": ""}


def format_amount(amount):
    if amount == '' or amount == 0 or amount == '0.000000000':
        return ''
    try:
        amount = abs(float(amount))  # Convertir a valor absoluto
    except ValueError:
        return amount
    if amount == int(amount):
        return str(int(amount))
    else:
        return f"{amount:.2f}"

def format_date_aaaammdd(date_value):
    if isinstance(date_value, date):
        return date_value.strftime("%Y%m%d")
    elif isinstance(date_value, str):
        return datetime.strptime(date_value, "%Y-%m-%d").strftime("%Y%m%d")
    return ''

def get_payment_methods(invoice_name):
    payment_entries = frappe.db.sql("""
        SELECT DISTINCT mop.custom_dgii_mode_of_payment
        FROM `tabPayment Entry Reference` per
        JOIN `tabPayment Entry` pe ON per.parent = pe.name
        JOIN `tabMode of Payment` mop ON pe.mode_of_payment = mop.name
        WHERE per.reference_name = %s
    """, (invoice_name,), as_dict=True)
        
    # Debugging output
    print("\n\n\n")
    print("Invoice Name:", invoice_name)
    print("Payment Entries:", payment_entries)
    
    if len(payment_entries) > 1:
        return '07'  # Mixta
    elif len(payment_entries) == 1:
        forma_de_pago = payment_entries[0].mode_of_payment
        return '01' if forma_de_pago == 'Efectivo' else '02' if forma_de_pago == 'Transferencia bancaria' else '03' if forma_de_pago == 'Tarjetas de credito' else ''
    return '04'

@frappe.whitelist()
def get_txt_file_address(from_date, to_date, decimal_places=2):
    # Obtener el tax_id del doctype Company
    company = frappe.get_doc("Company", frappe.defaults.get_user_default("Company"))
    rnc = company.tax_id.replace("-", "") if company.tax_id else ""

    # Calcular el periodo
    from_date_obj = datetime.strptime(from_date, "%Y-%m-%d")
    to_date_obj = datetime.strptime(to_date, "%Y-%m-%d")
    periodo = f"{from_date_obj.year}{min(from_date_obj.month, to_date_obj.month):02d}"

    # Consulta SQL para obtener los datos necesarios
    facturas = frappe.db.sql("""
        SELECT
            pinv.tax_id AS `RNC o Cedula`,
            CASE
                WHEN LENGTH(pinv.tax_id) = 9 THEN '1'
                WHEN LENGTH(pinv.tax_id) = 11 THEN '2'
                ELSE '3'
            END AS `Tipo Id`,
            pinv.tipo_bienes_y_servicios_comprados AS `Tipo Bienes y Servicios Comprados`,
            pinv.bill_no AS `NCF`,
            '' AS `NCF o Documento Modificado`,  # No disponible
            pinv.bill_date AS `Fecha Comprobante`,
            pe.reference_date AS `Fecha Pago`,
            pinv.monto_facturado_servicios AS `Monto Facturado en Servicios`,
            pinv.monto_facturado_bienes AS `Monto Facturado en Bienes`,
            pinv.base_total AS `Total Monto Facturado`,
            pinv.total_itbis AS `ITBIS Facturado`,
            pinv.retention_amount AS `ITBIS Retenido`,
            '' AS `ITBIS sujeto a Proporcionalidad (Art. 349)`,  # No disponible
            '' AS `ITBIS llevado al Costo`,  # No disponible
            '' AS `ITBIS por Adelantar`,  # No disponible
            '' AS `ITBIS percibido en compras`,  # No disponible
            pinv.retention_type AS `Tipo de Retencion en ISR`,
            pinv.isr_amount AS `Monto Retención Renta`,
            '' AS `ISR Percibido en compras`,  # No disponible
            pinv.excise_tax AS `Impuesto Selectivo al Consumo`,
            pinv.other_taxes AS `Otros Impuesto/Tasas`,
            pinv.legal_tip AS `Monto Propina Legal`,
            pinv.is_return AS `Es Nota de Débito`,
            pinv.return_against AS `Factura Original`,
            pinv.name AS `Factura Actual`
        FROM `tabPurchase Invoice` pinv
        LEFT JOIN `tabSupplier` supl ON supl.name = pinv.supplier
        LEFT JOIN `tabPayment Entry Reference` per ON per.reference_name = pinv.name
        LEFT JOIN `tabPayment Entry` pe ON pe.name = per.parent
        WHERE pinv.docstatus = 1
        AND pinv.bill_date BETWEEN %s AND %s
    """, (from_date, to_date), as_dict=True)

    # Calcular el número de registros
    numero_registros = len(facturas)

    # Crear el archivo TXT
    w = UnicodeWriter()

    # Agregar la primera línea
    w.writerow([
        "606", 
        rnc, 
        periodo, 
        numero_registros
    ])

    # Llenar las columnas del reporte 606 con los datos obtenidos
    for factura in facturas:
        ncf_modificado = ''
        # forma_de_pago = get_payment_methods(factura['Factura Original'] if factura['Es Nota de Débito'] else factura['Factura Actual'])
        forma_de_pago = get_payment_methods(factura['Factura Actual'])

        if factura['Es Nota de Débito']:
            original_invoice = frappe.get_doc("Purchase Invoice", factura['Factura Original'])
            ncf_modificado = original_invoice.bill_no
            forma_de_pago = "06" # "Nota de Crédito"

        tipo_bienes_y_servicios = factura['Tipo Bienes y Servicios Comprados']
        if tipo_bienes_y_servicios:
            tipo_bienes_y_servicios = tipo_bienes_y_servicios.split('-')[0]

        # Verificar si alguna columna de retención tiene un monto mayor a cero
        fecha_pago = ''
        if (factura['ITBIS Retenido'] and float(factura['ITBIS Retenido']) > 0) or \
           (factura['Monto Retención Renta'] and float(factura['Monto Retención Renta']) > 0):
            fecha_pago = format_date_aaaammdd(factura['Fecha Pago'])

        w.writerow([
            factura['RNC o Cedula'] or '',
            factura['Tipo Id'] or '',
            tipo_bienes_y_servicios or '',
            factura['NCF'] or '',
            ncf_modificado,
            format_date_aaaammdd(factura['Fecha Comprobante']),
            fecha_pago,
            format_amount(factura['Monto Facturado en Servicios']),
            format_amount(factura['Monto Facturado en Bienes']),
            format_amount(factura['Total Monto Facturado']),
            format_amount(factura['ITBIS Facturado']),
            format_amount(factura['ITBIS Retenido']),
            factura['ITBIS sujeto a Proporcionalidad (Art. 349)'] or '',
            factura['ITBIS llevado al Costo'] or '',
            factura['ITBIS por Adelantar'] or '',
            factura['ITBIS percibido en compras'] or '',
            factura['Tipo de Retencion en ISR'] or '',
            format_amount(factura['Monto Retención Renta']),
            factura['ISR Percibido en compras'] or '',
            format_amount(factura['Impuesto Selectivo al Consumo']),
            format_amount(factura['Otros Impuesto/Tasas']),
            format_amount(factura['Monto Propina Legal']),
            forma_de_pago
        ])

    # Convertir el contenido a texto con delimitador de pipes
    content = w.getvalue().replace(",", "|").replace('"', '')

    # Devolver el contenido del archivo de texto como descarga
    frappe.response['filename'] = "Reporte_606_%s.txt" % time.strftime("%Y%m%d_%H%M%S")
    frappe.response['filecontent'] = content
    frappe.response['type'] = 'download'


@frappe.whitelist()
def get_csv_file_address(from_date, to_date, decimal_places=2):
    # Obtener el tax_id del doctype Company
    company = frappe.get_doc("Company", frappe.defaults.get_user_default("Company"))
    rnc = company.tax_id.replace("-", "") if company.tax_id else ""

    # Calcular el periodo
    from_date_obj = datetime.strptime(from_date, "%Y-%m-%d")
    to_date_obj = datetime.strptime(to_date, "%Y-%m-%d")
    periodo = f"{from_date_obj.year}{min(from_date_obj.month, to_date_obj.month):02d}"

    # Consulta SQL para obtener los datos necesarios
    facturas = frappe.db.sql("""
        SELECT
            pinv.tax_id AS `RNC o Cedula`,
            CASE
                WHEN LENGTH(pinv.tax_id) = 9 THEN '1'
                WHEN LENGTH(pinv.tax_id) = 11 THEN '2'
                ELSE '3'
            END AS `Tipo Id`,
            pinv.tipo_bienes_y_servicios_comprados AS `Tipo Bienes y Servicios Comprados`,
            pinv.bill_no AS `NCF`,
            '' AS `NCF o Documento Modificado`,  # No disponible
            pinv.bill_date AS `Fecha Comprobante`,
            pe.reference_date AS `Fecha Pago`,
            pinv.monto_facturado_servicios AS `Monto Facturado en Servicios`,
            pinv.monto_facturado_bienes AS `Monto Facturado en Bienes`,
            pinv.base_total AS `Total Monto Facturado`,
            pinv.total_itbis AS `ITBIS Facturado`,
            pinv.retention_amount AS `ITBIS Retenido`,
            '' AS `ITBIS sujeto a Proporcionalidad (Art. 349)`,  # No disponible
            '' AS `ITBIS llevado al Costo`,  # No disponible
            '' AS `ITBIS por Adelantar`,  # No disponible
            '' AS `ITBIS percibido en compras`,  # No disponible
            pinv.retention_type AS `Tipo de Retencion en ISR`,
            pinv.isr_amount AS `Monto Retención Renta`,
            '' AS `ISR Percibido en compras`,  # No disponible
            pinv.excise_tax AS `Impuesto Selectivo al Consumo`,
            pinv.other_taxes AS `Otros Impuesto/Tasas`,
            pinv.legal_tip AS `Monto Propina Legal`,
            pinv.is_return AS `Es Nota de Débito`,
            pinv.return_against AS `Factura Original`,
            pinv.name AS `Factura Actual`
        FROM `tabPurchase Invoice` pinv
        LEFT JOIN `tabSupplier` supl ON supl.name = pinv.supplier
        LEFT JOIN `tabPayment Entry Reference` per ON per.reference_name = pinv.name
        LEFT JOIN `tabPayment Entry` pe ON pe.name = per.parent
        WHERE pinv.docstatus = 1
        AND pinv.bill_date BETWEEN %s AND %s
    """, (from_date, to_date), as_dict=True)

    # Calcular el número de registros
    numero_registros = len(facturas)
    print("\n\n\n")
    print("numero_registros:", numero_registros)

    # Crear el archivo CSV en memoria
    w = UnicodeWriter()

    # Agregar la primera línea
    w.writerow([
        "606", 
        rnc, 
        periodo, 
        numero_registros
    ])

    # Agregar el encabezado de las columnas
    w.writerow([
        "RNC o Cedula",
        "Tipo Id",
        "Tipo Bienes y Servicios Comprados",
        "NCF",
        "NCF o Documento Modificado",
        "Fecha Comprobante",
        "Fecha Pago",
        "Monto Facturado en Servicios",
        "Monto Facturado en Bienes",
        "Total Monto Facturado",
        "ITBIS Facturado",
        "ITBIS Retenido",
        "ITBIS sujeto a Proporcionalidad (Art. 349)",
        "ITBIS llevado al Costo",
        "ITBIS por Adelantar",
        "ITBIS percibido en compras",
        "Tipo de Retencion en ISR",
        "Monto Retención Renta",
        "ISR Percibido en compras",
        "Impuesto Selectivo al Consumo",
        "Otros Impuesto/Tasas",
        "Monto Propina Legal",
        "Forma de Pago"
    ])

    # Llenar las columnas del reporte 606 con los datos obtenidos
    for factura in facturas:
        ncf_modificado = ''
        forma_de_pago = get_payment_methods(factura['Factura Actual'])
        print(forma_de_pago)

        if factura['Es Nota de Débito']:
            original_invoice = frappe.get_doc("Purchase Invoice", factura['Factura Original'])
            ncf_modificado = original_invoice.bill_no
            forma_de_pago = "06" # "Nota de Crédito"

        tipo_bienes_y_servicios = factura['Tipo Bienes y Servicios Comprados']
        if tipo_bienes_y_servicios:
            tipo_bienes_y_servicios = tipo_bienes_y_servicios.split('-')[0]

        # Verificar si alguna columna de retención tiene un monto mayor a cero
        fecha_pago = ''
        if (factura['ITBIS Retenido'] and float(factura['ITBIS Retenido']) > 0) or \
           (factura['Monto Retención Renta'] and float(factura['Monto Retención Renta']) > 0):
            fecha_pago = format_date_aaaammdd(factura['Fecha Pago'])

        w.writerow([
            factura['RNC o Cedula'] or '',
            factura['Tipo Id'] or '',
            tipo_bienes_y_servicios or '',
            factura['NCF'] or '',
            ncf_modificado,
            format_date_aaaammdd(factura['Fecha Comprobante']),
            fecha_pago,
            format_amount(factura['Monto Facturado en Servicios']),
            format_amount(factura['Monto Facturado en Bienes']),
            format_amount(factura['Total Monto Facturado']),
            format_amount(factura['ITBIS Facturado']),
            format_amount(factura['ITBIS Retenido']),
            factura['ITBIS sujeto a Proporcionalidad (Art. 349)'] or '',
            factura['ITBIS llevado al Costo'] or '',
            factura['ITBIS por Adelantar'] or '',
            factura['ITBIS percibido en compras'] or '',
            factura['Tipo de Retencion en ISR'] or '',
            format_amount(factura['Monto Retención Renta']),
            factura['ISR Percibido en compras'] or '',
            format_amount(factura['Impuesto Selectivo al Consumo']),
            format_amount(factura['Otros Impuesto/Tasas']),
            format_amount(factura['Monto Propina Legal']),
            forma_de_pago
        ])

    # Devolver el contenido del archivo CSV como descarga
    frappe.response['result'] = cstr(w.getvalue())
    frappe.response['type'] = 'csv'
    frappe.response['doctype'] = "Reporte_606_" + str(int(time.time()))



def get_payment_method_id(invoice_name):
    payment_entries = frappe.db.sql("""
        SELECT DISTINCT mop.custom_dgii_mode_of_payment, pinv.outstanding_amount
        FROM `tabPayment Entry Reference` per
        JOIN `tabPayment Entry` pe ON per.parent = pe.name
        JOIN `tabPurchase Invoice` pinv ON per.reference_name = pinv.name
        JOIN `tabMode of Payment` mop ON pe.mode_of_payment = mop.name
        WHERE per.reference_name = %s AND pe.docstatus != 2
    """, (invoice_name,), as_dict=True)
    
    # Filtrar las entradas de pago con outstanding_amount == 0
    payment_entries = [entry for entry in payment_entries if entry.outstanding_amount == 0]
    
    if len(payment_entries) > 1:
        return '07 - MIXTA'  # Mixta
    elif len(payment_entries) == 1:
        forma_de_pago = payment_entries[0].custom_dgii_mode_of_payment
        return forma_de_pago
    return '04 - COMPRA A CREDITO'


@frappe.whitelist()
def get_excel_file_address(from_date, to_date, decimal_places=2):
    # Obtener el tax_id del doctype Company
    company = frappe.get_doc("Company", frappe.defaults.get_user_default("Company"))
    rnc = company.tax_id.replace("-", "") if company.tax_id else ""

    # Calcular el periodo
    from_date_obj = datetime.strptime(from_date, "%Y-%m-%d")
    to_date_obj = datetime.strptime(to_date, "%Y-%m-%d")
    periodo = f"{from_date_obj.year}{min(from_date_obj.month, to_date_obj.month):02d}"

    # Recuperar los valores de los campos del Doctype DGII Reports Settings
    settings = frappe.get_single("DGII Reports Settings")
    itbis_facturado = frappe.db.escape(settings.itbis_facturado or '')
    itbis_retenido = frappe.db.escape(settings.ret606_itbis_retenido or '')
    itbis_proporcionalidad = frappe.db.escape(settings.itbis_proporcionalidad or '')
    itbis_costo = frappe.db.escape(settings.itbis_costo or '')
    isc = frappe.db.escape(settings.isc or '')
    propina_legal = frappe.db.escape(settings.propina_legal or '')
    isr = frappe.db.escape(settings.ret606_isr or '')

    # Obtener las cuentas de la tabla otros_impuestos
    otros_impuestos_cuentas = [frappe.db.escape(item.cuenta) for item in settings.otros_impuestos]

    # Construir las condiciones SQL para otros_impuestos
    if otros_impuestos_cuentas:
        otros_impuestos_condition_ptc = " OR ".join([f"ptc.account_head = {cuenta}" for cuenta in otros_impuestos_cuentas])
        otros_impuestos_sql_ptc = f"SUM(CASE WHEN {otros_impuestos_condition_ptc} THEN ptc.tax_amount ELSE 0 END) AS `Otros Impuesto/Tasas`"
    else:
        otros_impuestos_sql_ptc = "0 AS `Otros Impuesto/Tasas`"

    # Subconsulta para obtener los otros impuestos
    otros_impuestos_subquery = f"""
        SELECT
            ptc.parent AS parent,
            {otros_impuestos_sql_ptc}
        FROM `tabPurchase Taxes and Charges` ptc
        GROUP BY ptc.parent
    """

    print("\n\n\n")
    print("otros_impuestos_subquery:", otros_impuestos_subquery)
    print("\n\n\n")
    
    # Consulta SQL para obtener los datos necesarios
    facturas = frappe.db.sql(f"""
        SELECT
            pinv.tax_id AS `RNC o Cedula`,
            CASE
                WHEN LENGTH(pinv.tax_id) = 9 THEN '1'
                WHEN LENGTH(pinv.tax_id) = 11 THEN '2'
                ELSE '3'
            END AS `Tipo Id`,
            pinv.tipo_bienes_y_servicios_comprados AS `Tipo Bienes y Servicios Comprados`,
            pinv.bill_no AS `NCF`,
            '' AS `NCF o Documento Modificado`,  # No disponible
            pinv.bill_date AS `Fecha Comprobante`,
            MAX(pe.reference_date) AS `Fecha Pago`,
            (SELECT COALESCE(SUM(pii.amount), 0)
            FROM `tabPurchase Invoice Item` pii
            JOIN `tabItem` item ON pii.item_code = item.item_code
            WHERE pii.parent = pinv.name AND item.item_type = 'Servicios') AS `Monto Facturado en Servicios`,
            (SELECT COALESCE(SUM(pii.amount), 0)
            FROM `tabPurchase Invoice Item` pii
            JOIN `tabItem` item ON pii.item_code = item.item_code
            WHERE pii.parent = pinv.name AND item.item_type = 'Bienes') AS `Monto Facturado en Bienes`,
            pinv.base_total AS `Total Monto Facturado`,
            (SELECT COALESCE(SUM(ptc.tax_amount), 0)
            FROM `tabPurchase Taxes and Charges` ptc
            WHERE ptc.parent = pinv.name AND ptc.account_head = {itbis_facturado}) AS `ITBIS Facturado`,
            (SELECT COALESCE(SUM(ptc.tax_amount), 0)
            FROM `tabPurchase Taxes and Charges` ptc
            WHERE ptc.parent = pinv.name AND ptc.account_head = {itbis_retenido}) AS `ITBIS Retenido`,
            (SELECT COALESCE(SUM(ptc.tax_amount), 0)
            FROM `tabPurchase Taxes and Charges` ptc
            WHERE ptc.parent = pinv.name AND ptc.account_head = {itbis_proporcionalidad}) AS `ITBIS sujeto a Proporcionalidad (Art. 349)`,
            (SELECT COALESCE(SUM(ptc.tax_amount), 0)
            FROM `tabPurchase Taxes and Charges` ptc
            WHERE ptc.parent = pinv.name AND ptc.account_head = {itbis_costo}) AS `ITBIS llevado al Costo`,
            '' AS `ITBIS por Adelantar`,  # No disponible
            '' AS `ITBIS percibido en compras`,  # No disponible
            pinv.retention_type AS `Tipo de Retencion en ISR`,
            (SELECT COALESCE(SUM(ptc.tax_amount), 0)
            FROM `tabPurchase Taxes and Charges` ptc
            WHERE ptc.parent = pinv.name AND ptc.account_head = {isr}) AS `Monto Retención Renta`,
            '' AS `ISR Percibido en compras`,  # No disponible
            (SELECT COALESCE(SUM(ptc.tax_amount), 0)
            FROM `tabPurchase Taxes and Charges` ptc
            WHERE ptc.parent = pinv.name AND ptc.account_head = {isc}) AS `Impuesto Selectivo al Consumo`,
            COALESCE(otros_impuestos.`Otros Impuesto/Tasas`, 0) AS `Otros Impuesto/Tasas`,
            (SELECT COALESCE(SUM(ptc.tax_amount), 0)
            FROM `tabPurchase Taxes and Charges` ptc
            WHERE ptc.parent = pinv.name AND ptc.account_head = {propina_legal}) AS `Monto Propina Legal`,
            pinv.is_return AS `Es Nota de Débito`,
            pinv.return_against AS `Factura Original`,
            pinv.name AS `Factura Actual`
        FROM `tabPurchase Invoice` pinv
        LEFT JOIN `tabSupplier` supl ON supl.name = pinv.supplier
        LEFT JOIN `tabPayment Entry Reference` per ON per.reference_name = pinv.name
        LEFT JOIN `tabPayment Entry` pe ON pe.name = per.parent
        LEFT JOIN ({otros_impuestos_subquery}) otros_impuestos ON otros_impuestos.parent = pinv.name
        WHERE pinv.docstatus = 1
          AND pinv.bill_date BETWEEN %s AND %s
        GROUP BY pinv.name

        UNION

        SELECT
            pinv.tax_id AS `RNC o Cedula`,
            CASE
                WHEN LENGTH(pinv.tax_id) = 9 THEN '1'
                WHEN LENGTH(pinv.tax_id) = 11 THEN '2'
                ELSE '3'
            END AS `Tipo Id`,
            pinv.tipo_bienes_y_servicios_comprados AS `Tipo Bienes y Servicios Comprados`,
            pinv.bill_no AS `NCF`,
            '' AS `NCF o Documento Modificado`,  # No disponible
            pinv.bill_date AS `Fecha Comprobante`,
            MAX(pe.reference_date) AS `Fecha Pago`,
            (SELECT COALESCE(SUM(pii.amount), 0)
            FROM `tabPurchase Invoice Item` pii
            JOIN `tabItem` item ON pii.item_code = item.item_code
            WHERE pii.parent = pinv.name AND item.item_type = 'Servicios') AS `Monto Facturado en Servicios`,
            (SELECT COALESCE(SUM(pii.amount), 0)
            FROM `tabPurchase Invoice Item` pii
            JOIN `tabItem` item ON pii.item_code = item.item_code
            WHERE pii.parent = pinv.name AND item.item_type = 'Bienes') AS `Monto Facturado en Bienes`,
            pinv.base_total AS `Total Monto Facturado`,
            (SELECT COALESCE(SUM(ptc2.tax_amount), 0)
            FROM `tabPurchase Taxes and Charges` ptc2
            WHERE ptc2.parent = pinv.name AND ptc2.account_head = {itbis_facturado}) AS `ITBIS Facturado`,
            (SELECT COALESCE(SUM(ptc2.tax_amount), 0)
            FROM `tabPurchase Taxes and Charges` ptc2
            WHERE ptc2.parent = pinv.name AND ptc2.account_head = {itbis_retenido}) AS `ITBIS Retenido`,
            (SELECT COALESCE(SUM(ptc2.tax_amount), 0)
            FROM `tabPurchase Taxes and Charges` ptc2
            WHERE ptc2.parent = pinv.name AND ptc2.account_head = {itbis_proporcionalidad}) AS `ITBIS sujeto a Proporcionalidad (Art. 349)`,
            (SELECT COALESCE(SUM(ptc2.tax_amount), 0)
            FROM `tabPurchase Taxes and Charges` ptc2
            WHERE ptc2.parent = pinv.name AND ptc2.account_head = {itbis_costo}) AS `ITBIS llevado al Costo`,
            '' AS `ITBIS por Adelantar`,  # No disponible
            '' AS `ITBIS percibido en compras`,  # No disponible
            pinv.retention_type AS `Tipo de Retencion en ISR`,
            (SELECT COALESCE(SUM(ptc2.tax_amount), 0)
            FROM `tabPurchase Taxes and Charges` ptc2
            WHERE ptc2.parent = pinv.name AND ptc2.account_head = {isr}) AS `Monto Retención Renta`,
            '' AS `ISR Percibido en compras`,  # No disponible
            (SELECT COALESCE(SUM(ptc2.tax_amount), 0)
            FROM `tabPurchase Taxes and Charges` ptc2
            WHERE ptc2.parent = pinv.name AND ptc2.account_head = {isc}) AS `Impuesto Selectivo al Consumo`,
            COALESCE(otros_impuestos.`Otros Impuesto/Tasas`, 0) AS `Otros Impuesto/Tasas`,
            (SELECT COALESCE(SUM(ptc2.tax_amount), 0)
            FROM `tabPurchase Taxes and Charges` ptc2
            WHERE ptc2.parent = pinv.name AND ptc2.account_head = {propina_legal}) AS `Monto Propina Legal`,
            pinv.is_return AS `Es Nota de Débito`,
            pinv.return_against AS `Factura Original`,
            pinv.name AS `Factura Actual`
        FROM `tabPurchase Invoice` pinv
        LEFT JOIN `tabSupplier` supl ON supl.name = pinv.supplier
        LEFT JOIN `tabPayment Entry Reference` per ON per.reference_name = pinv.name
        LEFT JOIN `tabPayment Entry` pe ON pe.name = per.parent
        LEFT JOIN ({otros_impuestos_subquery}) otros_impuestos ON otros_impuestos.parent = pinv.name
        WHERE pinv.docstatus = 1
          AND pe.reference_date BETWEEN %s AND %s
          AND pinv.bill_date < %s
          AND pinv.outstanding_amount = 0
        GROUP BY pinv.name
        ORDER BY `Fecha Comprobante`, `Factura Actual`
    """, (from_date, to_date, from_date, to_date, from_date), as_dict=True)

    # Calcular el número de registros
    numero_registros = len(facturas)
    
    # Crear el archivo Excel en memoria
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte 606"

    # Estilos
    header_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")  # Verde oscuro
    header_font = Font(color="FFFFFF", bold=True)
    header_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    detail_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    text_style = NamedStyle(name="text_style", number_format="@")

    # Agregar la primera línea
    ws.merge_cells('A1:B1')
    ws['A1'] = "RNC o Cédula"
    ws['C1'] = rnc
    ws['A1'].fill = header_fill
    ws['A1'].font = header_font
    ws['A1'].border = header_border
    ws['B1'].border = header_border
    ws['C1'].border = header_border  # Aplicar borde a la celda C1
    ws['A1'].alignment = Alignment(horizontal='right')
    ws['C1'].alignment = Alignment(horizontal='right')

    # Agregar la segunda línea
    ws.merge_cells('A2:B2')
    ws['A2'] = "Periodo"
    ws['C2'] = periodo
    ws['A2'].fill = header_fill
    ws['A2'].font = header_font
    ws['A2'].border = header_border
    ws['B2'].border = header_border
    ws['C2'].border = header_border
    ws['A2'].alignment = Alignment(horizontal='right')
    ws['C2'].alignment = Alignment(horizontal='right')

    # Agregar la tercera línea
    ws.merge_cells('A3:B3')
    ws['A3'] = "Cantidad Registros"
    ws['C3'] = numero_registros
    ws['A3'].fill = header_fill
    ws['A3'].font = header_font
    ws['A3'].border = header_border
    ws['B3'].border = header_border
    ws['C3'].border = header_border
    ws['A3'].alignment = Alignment(horizontal='right')
    ws['C3'].alignment = Alignment(horizontal='right')

    # Dejar un renglón en blanco
    ws.append([])

    # Agregar el encabezado de detalles
    ws.merge_cells('B5:Z5')  # Combinar hasta la columna Z5
    ws['B5'] = "Detalles"
    ws['B5'].alignment = Alignment(horizontal='center')
    ws['B5'].fill = header_fill
    ws['B5'].font = header_font
    ws['B5'].border = header_border

    # Agregar la fila de números de columna
    column_numbers = ["",1,2,3,4,5,6,6,7,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23]
    ws.append(column_numbers)
    for col_num in range(2, 27):
        cell = ws.cell(row=6, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = Alignment(horizontal='center')  # Centrar el texto

    # Combinar celdas para las columnas Fecha Comprobante y Fecha Pago en la numeración
    ws.merge_cells('G6:H6')
    ws.merge_cells('I6:J6')

    # Agregar el encabezado de las columnas
    headers = [
        "Líneas",
        "RNC o Cedula",
        "Tipo Id",
        "Tipo Bienes y Servicios Comprados",
        "NCF",
        "NCF o Documento Modificado",
        "Fecha Comprobante",
        "",
        "Fecha Pago",
        "",
        "Monto Facturado en Servicios",
        "Monto Facturado en Bienes",
        "Total Monto Facturado",
        "ITBIS Facturado",
        "ITBIS Retenido",
        "ITBIS sujeto a Proporcionalidad (Art. 349)",
        "ITBIS llevado al Costo",
        "ITBIS por Adelantar",
        "ITBIS percibido en compras",
        "Tipo de Retencion en ISR",
        "Monto Retención Renta",
        "ISR Percibido en compras",
        "Impuesto Selectivo al Consumo",
        "Otros Impuesto/Tasas",
        "Monto Propina Legal",
        "Forma de Pago"
    ]
    ws.append(headers)
    for col_num in range(1, 27):
        cell = ws.cell(row=7, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)  # Centrar y ajustar el texto

    # Combinar celdas para las columnas Fecha Comprobante y Fecha Pago en el encabezado
    ws.merge_cells('G7:H7')
    ws.merge_cells('I7:J7')

    # Ajustar el ancho de las columnas
    column_widths = {
        'A': 10,
        'B': 15,
        'C': 10,
        'D': 50,
        'E': 15,
        'F': 15,
        'G': 8,
        'H': 5,
        'I': 8,
        'J': 5,
        'K': 15,
        'L': 15,
        'M': 15,
        'N': 15,
        'O': 15,
        'P': 15,
        'Q': 15,
        'R': 15,
        'S': 15,
        'T': 15,
        'U': 15,
        'V': 15,
        'W': 15,
        'X': 15,
        'Y': 15,
        'Z': 15
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Llenar las columnas del reporte 606 con los datos obtenidos
    for idx, factura in enumerate(facturas, start=1):
        ncf_modificado = ''
        forma_de_pago = get_payment_method_id(factura['Factura Actual'])
        if factura['Es Nota de Débito']:
            original_invoice = frappe.get_doc("Purchase Invoice", factura['Factura Original'])
            ncf_modificado = original_invoice.bill_no
            forma_de_pago = "06 - NOTA DE CREDITO"  # "Nota de Crédito"

        tipo_bienes_y_servicios = factura['Tipo Bienes y Servicios Comprados']

        # Verificar si alguna columna de retención tiene un monto mayor a cero
        fecha_pago = ''
        if (factura['ITBIS Retenido'] and float(factura['ITBIS Retenido']) > 0) or \
        (factura['Monto Retención Renta'] and float(factura['Monto Retención Renta']) > 0):
            fecha_pago = factura['Fecha Pago']

        # Separar las fechas en AAAAMM y DD
        fecha_comprobante = factura['Fecha Comprobante']
        fecha_comprobante_aaaamm = fecha_comprobante.strftime("%Y%m")
        fecha_comprobante_dd = fecha_comprobante.strftime("%d")

        fecha_pago_aaaamm = ''
        fecha_pago_dd = ''
        if fecha_pago:
            fecha_pago_aaaamm = fecha_pago.strftime("%Y%m")
            fecha_pago_dd = fecha_pago.strftime("%d")

        # Calcular el número de pagos
        num_pagos = frappe.db.count('Payment Entry Reference', {'reference_name': factura['Factura Actual']})

        # Dividir cada impuesto entre el número de pagos si es mayor o igual a dos
        divisor = num_pagos if num_pagos >= 2 else 1
        divisor = 1

        row = [
            idx,
            factura['RNC o Cedula'] or '',
            factura['Tipo Id'] or '',
            tipo_bienes_y_servicios or '',
            factura['NCF'] or '',
            ncf_modificado,
            fecha_comprobante_aaaamm,
            fecha_comprobante_dd,
            fecha_pago_aaaamm,
            fecha_pago_dd,
            format_amount(factura['Monto Facturado en Servicios'] ),
            format_amount(factura['Monto Facturado en Bienes'] ),
            format_amount(factura['Total Monto Facturado'] ),
            format_amount(factura['ITBIS Facturado'] / divisor),
            format_amount(factura['ITBIS Retenido'] / divisor),
            factura['ITBIS sujeto a Proporcionalidad (Art. 349)'] or '',
            factura['ITBIS llevado al Costo'] or '',
            factura['ITBIS por Adelantar'] or '',
            factura['ITBIS percibido en compras'] or '',
            factura['Tipo de Retencion en ISR'] or '',
            format_amount(factura['Monto Retención Renta'] / divisor),
            factura['ISR Percibido en compras'] or '',
            format_amount(factura['Impuesto Selectivo al Consumo'] / divisor),
            format_amount(factura['Otros Impuesto/Tasas'] / divisor),
            format_amount(factura['Monto Propina Legal'] / divisor),
            forma_de_pago
        ]
        ws.append(row)
        if "text_style" not in wb.named_styles:
            wb.add_named_style(text_style)
        for col_num in range(1, 27):
            cell = ws.cell(row=idx + 7, column=col_num)
            cell.style = text_style  # Aplicar el estilo de texto plano
            if col_num == 1:
                cell.fill = detail_fill
                cell.alignment = Alignment(horizontal='center')  # Centrar el texto horizontalmente
    # Guardar el archivo Excel en memoria
    wb.save(output)
    output.seek(0)

    # Devolver el contenido del archivo Excel como descarga
    frappe.response['filename'] = "Reporte_606_%s.xlsx" % time.strftime("%Y%m%d_%H%M%S")
    frappe.response['filecontent'] = output.getvalue()
    frappe.response['type'] = 'download'


# MariaDB [_d31ecb88ac1f70bf]> SELECT pinv.tax_id, pinv.tipo_bienes_y_servicios_comprados FROM `tabPurchase Invoice` pinv INNER JOIN `tabSupplier` supl ON supl.name = pinv.suppl
#  Charges` ptc ON ptc.parent = pinv.name WHERE pinv.docstatus = 1;
# +-----------+-------------------------------------------------+
# | tax_id    | tipo_bienes_y_servicios_comprados               |
# +-----------+-------------------------------------------------+
# | 111000475 | 02-GASTOS POR TRABAJOS, SUMINISTROS Y SERVICIOS |
# | 111000475 | 02-GASTOS POR TRABAJOS, SUMINISTROS Y SERVICIOS |
# | 111000475 | 02-GASTOS POR TRABAJOS, SUMINISTROS Y SERVICIOS |
# | 111000475 | 02-GASTOS POR TRABAJOS, SUMINISTROS Y SERVICIOS |
# | 111000475 | 02-GASTOS POR TRABAJOS, SUMINISTROS Y SERVICIOS |
# | 111000475 | 02-GASTOS POR TRABAJOS, SUMINISTROS Y SERVICIOS |
# +-----------+-------------------------------------------------+
# 6 rows in set (0.007 sec)

# MariaDB [_d31ecb88ac1f70bf]> SELECT pinv.tax_id, pinv.tipo_bienes_y_servicios_comprados FROM `tabPurchase Invoice` pinv INNER JOIN `tabSupplier` supl ON supl.name = pinv.supplier LEFT JOIN `tabPayment Entry Reference` per ON per.reference_name = pinv.name LEFT JOIN `tabPayment Entry` pe ON pe.name = per.parent WHERE pinv.docstatus = 1;
# +-----------+-------------------------------------------------+
# | tax_id    | tipo_bienes_y_servicios_comprados               |
# +-----------+-------------------------------------------------+
# | 111000475 | 02-GASTOS POR TRABAJOS, SUMINISTROS Y SERVICIOS |
# | 111000475 | 02-GASTOS POR TRABAJOS, SUMINISTROS Y SERVICIOS |
# +-----------+-------------------------------------------------+
# 2 rows in set (0.025 sec)

# MariaDB [_d31ecb88ac1f70bf]> SELECT pinv.tax_id, pinv.tipo_bienes_y_servicios_comprados FROM `tabPurchase Invoice` pinv INNER JOIN `tabSupplier` supl ON supl.name = pinv.supplier LEFT JOIN `tabPayment Entry Reference` per ON per.reference_name = pinv.name WHERE pinv.docstatus = 1;
# +-----------+-------------------------------------------------+
# | tax_id    | tipo_bienes_y_servicios_comprados               |
# +-----------+-------------------------------------------------+
# | 111000475 | 02-GASTOS POR TRABAJOS, SUMINISTROS Y SERVICIOS |
# | 111000475 | 02-GASTOS POR TRABAJOS, SUMINISTROS Y SERVICIOS |
# +-----------+-------------------------------------------------+
# 2 rows in set (0.015 sec)

# MariaDB [_d31ecb88ac1f70bf]> SELECT pinv.tax_id, pinv.tipo_bienes_y_servicios_comprados FROM `tabPurchase Invoice` pinv INNER JOIN `tabSupplier` supl ON supl.name = pinv.supplier WHERE pinv.docstatus = 1;
# +-----------+-------------------------------------------------+
# | tax_id    | tipo_bienes_y_servicios_comprados               |
# +-----------+-------------------------------------------------+
# | 111000475 | 02-GASTOS POR TRABAJOS, SUMINISTROS Y SERVICIOS |
# +-----------+-------------------------------------------------+
# 1 row in set (0.001 sec)

# MariaDB [_d31ecb88ac1f70bf]>