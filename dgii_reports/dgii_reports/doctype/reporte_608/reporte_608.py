# Copyright (c) 2024, TnologiaRD and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, cint, flt, formatdate, format_datetime
from frappe.model.document import Document
from frappe.utils.csvutils import UnicodeWriter
import time
from datetime import datetime
from frappe import _
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side, NamedStyle

class Reporte608(Document):
	pass

@frappe.whitelist()
def validate_pending_invoices(from_date, to_date):
    # Consulta para facturas de venta
    draft_sales_invoices = frappe.db.sql("""
        SELECT 
            COUNT(*) 
        FROM 
            `tabSales Invoice` 
        WHERE 
            docstatus = 0 AND posting_date BETWEEN '%s' AND '%s'
    """ % (from_date, to_date))

    # Consulta para facturas de compra
    draft_purchase_invoices = frappe.db.sql("""
        SELECT 
            COUNT(*) 
        FROM 
            `tabPurchase Invoice` 
        WHERE 
            docstatus = 0 AND posting_date BETWEEN '%s' AND '%s'
    """ % (from_date, to_date))

    # Sumar los resultados de ambas consultas
    total_draft_invoices = draft_sales_invoices[0][0] + draft_purchase_invoices[0][0]

    if total_draft_invoices > 0:
        frappe.log_error("Facturas pendientes encontradas", "validate_pending_invoices")
        return {"message": "Hay facturas de venta o compra pendientes de procesar. Por favor, complete las facturas pendientes antes de generar el reporte."}
    
    frappe.log_error("No hay facturas pendientes", "validate_pending_invoices")
    return {"message": ""}



@frappe.whitelist()
def get_file_1__address(from_date, to_date, decimal_places=2):
    # Obtener el tax_id del doctype Company
    company = frappe.get_doc("Company", frappe.defaults.get_user_default("Company"))
    rnc = company.tax_id.replace("-", "") if company.tax_id else ""

    # Calcular el periodo
    from_date_obj = datetime.strptime(from_date, "%Y-%m-%d")
    to_date_obj = datetime.strptime(to_date, "%Y-%m-%d")
    periodo = f"{from_date_obj.year}{min(from_date_obj.month, to_date_obj.month):02d}"

    # Consulta SQL para obtener los datos necesarios
    query = """
        SELECT 
            sinv.custom_ncf as ncf, 
            sinv.posting_date as fecha_comprobante,
            sinv.custom_codigo_de_anulacion as tipo_anulacion  -- Recuperar la columna codigo de la tabla tabTipo de Anulacion
        FROM 
            `tabSales Invoice` AS sinv 
        WHERE 
            sinv.docstatus = 2 AND sinv.posting_date BETWEEN %(from_date)s AND %(to_date)s

        UNION

        SELECT 
            pinv.bill_no as ncf, 
            pinv.posting_date as fecha_comprobante,
            pinv.custom_codigo_de_anulacion as tipo_anulacion  -- Recuperar la columna codigo de la tabla tabTipo de Anulacion
        FROM 
            `tabPurchase Invoice` AS pinv 
        WHERE 
            pinv.docstatus = 2 AND pinv.posting_date BETWEEN %(from_date)s AND %(to_date)s
            AND (pinv.bill_no LIKE 'B13%%' OR pinv.bill_no LIKE 'B11%%')
    """
    # Ejecutar la consulta
    result = frappe.db.sql(query, {"from_date": from_date, "to_date": to_date}, as_dict=True)    # Número de registros
    numero_registros = len(result)

    # Crear el archivo en memoria usando UnicodeWriter
    w = UnicodeWriter()

    # Agregar la primera línea
    w.writerow([
        "608", 
        rnc, 
        periodo, 
        numero_registros
    ])

    for row in result:
        w.writerow([
            row.ncf, 
            row.fecha_comprobante.strftime("%Y%m%d") if row.fecha_comprobante else "", 
            row.tipo_anulacion
        ])

    # Convertir el contenido a texto con delimitador de pipes
    content = w.getvalue().replace(",", "|").replace('"', '')

    # Devolver el contenido del archivo de texto como descarga
    frappe.response['filename'] = "Reporte_607_%s.txt" % time.strftime("%Y%m%d_%H%M%S")
    frappe.response['filecontent'] = content
    frappe.response['type'] = 'download'


@frappe.whitelist()
def get_file_address(from_date, to_date, decimal_places=2):
    # Obtener el tax_id del doctype Company
    company = frappe.get_doc("Company", frappe.defaults.get_user_default("Company"))
    rnc = company.tax_id.replace("-", "") if company.tax_id else ""

    # Calcular el periodo
    from_date_obj = datetime.strptime(from_date, "%Y-%m-%d")
    to_date_obj = datetime.strptime(to_date, "%Y-%m-%d")
    periodo = f"{from_date_obj.year}{min(from_date_obj.month, to_date_obj.month):02d}"

    # Consulta SQL para obtener los datos necesarios
    query = """
        SELECT 
            sinv.custom_ncf as ncf, 
            sinv.posting_date as fecha_comprobante,
            sinv.custom_codigo_de_anulacion as tipo_anulacion  -- Recuperar la columna codigo de la tabla tabTipo de Anulacion
        FROM 
            `tabSales Invoice` AS sinv 
        WHERE 
            sinv.docstatus = 2 
            AND sinv.posting_date BETWEEN %(from_date)s AND %(to_date)s
            AND sinv.name NOT IN (SELECT amended_from FROM `tabSales Invoice` WHERE amended_from IS NOT NULL)

        UNION

        SELECT 
            pinv.bill_no as ncf, 
            pinv.posting_date as fecha_comprobante,
            pinv.custom_codigo_de_anulacion as tipo_anulacion  -- Recuperar la columna codigo de la tabla tabTipo de Anulacion
        FROM 
            `tabPurchase Invoice` AS pinv 
        WHERE 
            pinv.docstatus = 2 
            AND pinv.posting_date BETWEEN %(from_date)s AND %(to_date)s
            AND (pinv.bill_no LIKE 'B13%%' OR pinv.bill_no LIKE 'B14%%')
            AND pinv.name NOT IN (SELECT amended_from FROM `tabPurchase Invoice` WHERE amended_from IS NOT NULL)
    """
    # Ejecutar la consulta
    facturas = frappe.db.sql(query, {"from_date": from_date, "to_date": to_date}, as_dict=True)

    # Número de registros
    numero_registros = len(facturas)

    # Crear el archivo Excel en memoria
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte 608S"

    # Estilos
    header_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")  # Verde oscuro
    header_font = Font(color="FFFFFF", bold=True)
    header_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    detail_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    text_style = NamedStyle(name="text_style", number_format="@")

    # Agregar la primera línea
    ws.merge_cells('A1:B1')
    ws['A1'] = "RNC o Cédula"
    ws['C1'] = rnc
    ws['A1'].fill = header_fill
    ws['A1'].font = header_font
    ws['A1'].border = header_border
    ws['B1'].border = header_border
    ws['C1'].border = header_border  # Aplicar borde a la celda C1
    ws['A1'].alignment = Alignment(horizontal='right')
    ws['C1'].alignment = Alignment(horizontal='right')
    # Agregar la segunda línea
    ws.merge_cells('A2:B2')
    ws['A2'] = "Periodo"
    ws['C2'] = periodo
    ws['A2'].fill = header_fill
    ws['A2'].font = header_font
    ws['A2'].border = header_border
    ws['B2'].border = header_border
    ws['C2'].border = header_border
    ws['A2'].alignment = Alignment(horizontal='right')
    ws['C2'].alignment = Alignment(horizontal='right')
    # Agregar la tercera línea
    ws.merge_cells('A3:B3')
    ws['A3'] = "Cantidad Registros"
    ws['C3'] = numero_registros
    ws['A3'].fill = header_fill
    ws['A3'].font = header_font
    ws['A3'].border = header_border
    ws['B3'].border = header_border
    ws['C3'].border = header_border
    ws['A3'].alignment = Alignment(horizontal='right')
    ws['C3'].alignment = Alignment(horizontal='right')
    # Dejar un renglón en blanco
    ws.append([])

    # Agregar el encabezado de detalles
    ws.merge_cells('B5:F5')  # Combinar hasta la columna F5
    ws['B5'] = "Detalles"
    ws['B5'].alignment = Alignment(horizontal='center')
    ws['B5'].fill = header_fill
    ws['B5'].font = header_font
    ws['B5'].border = header_border

    # Agregar la fila de números de columna
    column_numbers = ["", 1, 1, 2, 3, 3]
    ws.append(column_numbers)
    for col_num in range(2, 6):
        cell = ws.cell(row=6, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = Alignment(horizontal='center')  # Centrar el texto

    # Combinar celdas para las columnas Fecha Comprobante y Fecha Pago en la numeración
    ws.merge_cells('B6:C6')
    ws.merge_cells('E6:F6')

    # Agregar el encabezado de las columnas
    headers = [
        "Líneas",
        "Número de Comprobante Fiscal",
        "",
        "Fecha de Comprobante",
        "Tipo de Anulación",
        ""
    ]
    ws.append(headers)
    for col_num in range(1, 6):
        cell = ws.cell(row=7, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Centrar y ajustar el texto

    # Combinar celdas para las columnas Fecha Comprobante y Fecha Pago en el encabezado
    ws.merge_cells('B7:C7')
    ws.merge_cells('E7:F7')

    # Ajustar el ancho de las columnas
    column_widths = {
        'A': 10,
        'B': 20,
        'C': 20,
        'D': 20,
        'E': 20,
        'F': 20,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    for idx, factura in enumerate(facturas, start=1):
        fecha = factura['fecha_comprobante'].strftime("%Y%m%d") if factura['fecha_comprobante'] else ""
        tipo_anulacion_codigo = factura['tipo_anulacion']

        # Recuperar el doctype Tipo de Anulacion usando el código
        try:
            tipo_anulacion_doc = frappe.db.get_value("Tipo de Anulacion", {"codigo": tipo_anulacion_codigo}, ["tipo_de_anulacion", "codigo"])
            if tipo_anulacion_doc:
                tipo_anulacion = f"{tipo_anulacion_doc[1]} {tipo_anulacion_doc[0]}"
            else:
                tipo_anulacion = f"Desconocido - {tipo_anulacion_codigo}"
        except Exception as e:
            tipo_anulacion = f"Error - {tipo_anulacion_codigo}"
            frappe.log_error(message=str(e), title="Error al recuperar Tipo de Anulacion")
        
        row = [
            idx,
            factura['ncf'] or '',
            '',  # Celda vacía para la combinación
            fecha or '',
            tipo_anulacion or '',
            ''  # Celda vacía para la combinación
        ]
        ws.append(row)
        if "text_style" not in wb.named_styles:
            wb.add_named_style(text_style)
        for col_num in range(1, 7):
            cell = ws.cell(row=idx + 7, column=col_num)
            cell.style = text_style  # Aplicar el estilo de texto plano
            cell.alignment = Alignment(horizontal='center', vertical='center')  # Centrar el texto horizontal y verticalmente
            if col_num == 1:
                cell.fill = detail_fill

        # Combinar celdas para el segundo valor (B y C)
        ws.merge_cells(start_row=idx + 7, start_column=2, end_row=idx + 7, end_column=3)
        # Combinar celdas para el cuarto valor (E y F)
        ws.merge_cells(start_row=idx + 7, start_column=5, end_row=idx + 7, end_column=6)

    # Guardar el archivo Excel en memoria
    wb.save(output)
    output.seek(0)

    # Devolver el contenido del archivo Excel como descarga
    frappe.response['filename'] = "Reporte_608_%s.xlsx" % time.strftime("%Y%m%d_%H%M%S")
    frappe.response['filecontent'] = output.getvalue()
    frappe.response['type'] = 'download'